"""
FastAPI server for SAP SF Permission Comparer.

Serves the frontend and provides API endpoints for:
  - Comparing T3 vs PROD permission PDFs
  - Comparing a PDF role export against an Excel workbook
  - Generating an Excel file from a PDF permission export
"""

import io
import os
import tempfile

import openpyxl
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse

from comparer import (
    compare_permissions,
    compare_raw_lines,
    extract_pdf_structured,
    extract_text_from_pdf,
    parse_permission_lines,
    parse_sections,
)
from compare_pdf_vs_excel import (
    build_excel_lookup,
    compare_pdf_vs_excel,
    extract_excel_permissions,
    extract_pdf_permissions,
    generate_pdf_excel,
)

app = FastAPI(title="SAP SF Permission Comparer")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

FRONTEND_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "frontend")


@app.get("/")
async def serve_frontend():
    """Serve the frontend single-page application."""
    return FileResponse(os.path.join(FRONTEND_DIR, "index.html"))


@app.get("/api/info")
async def info():
    """Return build metadata."""
    return {"buildDate": os.environ.get("BUILD_DATE", "unknown")}


@app.post("/api/compare")
async def compare(t3: UploadFile = File(...), prod: UploadFile = File(...)):
    """
    Accept two PDF uploads (t3 and prod), run the comparison pipeline,
    and return the structured differences and raw diff as JSON.
    """
    # Write uploads to temp files (PyMuPDF needs file paths)
    t3_tmp = None
    prod_tmp = None
    try:
        t3_tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        t3_tmp.write(await t3.read())
        t3_tmp.close()

        prod_tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        prod_tmp.write(await prod.read())
        prod_tmp.close()

        # Extract text
        t3_text = extract_text_from_pdf(t3_tmp.name)
        prod_text = extract_text_from_pdf(prod_tmp.name)

        # Parse sections
        t3_sections = parse_sections(t3_text)
        prod_sections = parse_sections(prod_text)

        # Compare
        differences = compare_permissions(t3_sections, prod_sections)
        raw_diff = compare_raw_lines(t3_text, prod_text)

        # Normalize difference keys to camelCase for the frontend
        normalized_differences = []
        for diff in differences:
            normalized_differences.append({
                "section": diff["section"],
                "item": diff["item"],
                "type": diff["type"],
                "t3Value": diff["t3_value"],
                "prodValue": diff["prod_value"],
            })

        return {
            "differences": normalized_differences,
            "rawDiff": {
                "onlyInT3": raw_diff["only_in_t3"],
                "onlyInProd": raw_diff["only_in_prod"],
            },
            "t3FileName": t3.filename,
            "prodFileName": prod.filename,
        }

    finally:
        # Clean up temp files
        if t3_tmp and os.path.exists(t3_tmp.name):
            os.unlink(t3_tmp.name)
        if prod_tmp and os.path.exists(prod_tmp.name):
            os.unlink(prod_tmp.name)


@app.post("/api/compare-pdf-excel")
async def compare_pdf_excel(pdf: UploadFile = File(...), excel: UploadFile = File(...)):
    """
    Accept a PDF role export and an Excel workbook, compare permissions,
    and return a structured JSON report.

    The role name is detected from the PDF filename and used to locate
    the matching column in the "ROLE ACCESS (WHAT)" Excel sheet.
    """
    pdf_tmp = None
    excel_tmp = None
    try:
        # Write uploads to temp files
        pdf_tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        pdf_tmp.write(await pdf.read())
        pdf_tmp.close()

        excel_tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        excel_tmp.write(await excel.read())
        excel_tmp.close()

        # Extract PDF permissions
        try:
            pdf_entries = extract_pdf_permissions(pdf_tmp.name)
        except Exception as exc:
            raise HTTPException(status_code=422, detail=f"PDF parsing error: {exc}") from exc

        if not pdf_entries:
            raise HTTPException(status_code=422, detail="No permission entries found in PDF.")

        # Detect role name from PDF filename (e.g. "View Role for ALL_MGR_GL_Manager NEW.pdf")
        pdf_basename = os.path.splitext(pdf.filename or "")[0]
        role_name = None
        import re as _re
        # Try "View Role for <role_name>" pattern
        m = _re.search(r'View Role for\s+(.+)', pdf_basename, _re.IGNORECASE)
        if m:
            role_name = m.group(1).strip()
        # Fall back to the full basename
        if not role_name:
            role_name = pdf_basename.strip()

        EXCEL_SHEET = "ROLE ACCESS (WHAT)"

        # Open Excel and find the column matching the role name
        try:
            wb = openpyxl.load_workbook(excel_tmp.name, data_only=True)
        except Exception as exc:
            raise HTTPException(status_code=422, detail=f"Excel open error: {exc}") from exc

        if EXCEL_SHEET not in wb.sheetnames:
            raise HTTPException(
                status_code=422,
                detail=f"Sheet '{EXCEL_SHEET}' not found in Excel workbook. "
                       f"Available sheets: {wb.sheetnames}",
            )

        ws = wb[EXCEL_SHEET]
        # Scan header rows (rows 1-3) to find the role column
        role_col = None
        role_name_norm = role_name.lower().strip()
        for header_row in range(1, 4):
            for col in range(1, ws.max_column + 1):
                cell_val = ws.cell(row=header_row, column=col).value
                if cell_val and role_name_norm in str(cell_val).lower().strip():
                    role_col = col
                    break
            if role_col:
                break
        wb.close()

        if role_col is None:
            raise HTTPException(
                status_code=422,
                detail=f"Role '{role_name}' not found in header rows of sheet '{EXCEL_SHEET}'. "
                       "Make sure the PDF filename contains the exact role name as it appears "
                       "in the Excel column header.",
            )

        # Extract Excel permissions using the detected column
        try:
            excel_entries = extract_excel_permissions(excel_tmp.name, EXCEL_SHEET, role_col)
        except Exception as exc:
            raise HTTPException(status_code=422, detail=f"Excel parsing error: {exc}") from exc

        # Compare
        missing_in_excel, missing_in_pdf, mismatches, matched = compare_pdf_vs_excel(
            pdf_entries, excel_entries
        )

        def _fmt_pdf_entry(e):
            return {
                "section": e.get("element") or e.get("section", ""),
                "subsection": e.get("subsection", ""),
                "field": e.get("field", ""),
                "permissions": e.get("permissions_str", ""),
                "excelRow": e.get("_excel_row"),
                "excelHasField": e.get("_excel_has_field", False),
                "excelValue": e.get("_excel_value"),
            }

        def _fmt_excel_entry(e):
            return {
                "element": e.get("element", ""),
                "grouping": e.get("grouping", ""),
                "permission": e.get("permission", ""),
                "value": e.get("value", ""),
                "row": e.get("row"),
            }

        truly_missing = [e for e in missing_in_excel if not e.get("_excel_has_field")]
        has_field_none = [e for e in missing_in_excel if e.get("_excel_has_field")]

        return {
            "pdfFile": pdf.filename,
            "excelFile": excel.filename,
            "roleName": role_name,
            "roleColumn": role_col,
            "summary": {
                "totalPdfPerms": len(pdf_entries),
                "totalExcelPerms": len(excel_entries),
                "matched": len(matched),
                "onlyInPdf": len(truly_missing),
                "excelFieldNone": len(has_field_none),
                "onlyInExcel": len(missing_in_pdf),
                "mismatches": len(mismatches),
            },
            "onlyInPdf": [_fmt_pdf_entry(e) for e in truly_missing],
            "excelFieldNone": [_fmt_pdf_entry(e) for e in has_field_none],
            "onlyInExcel": [_fmt_excel_entry(e) for e in missing_in_pdf],
            "mismatches": [
                {
                    "section": (m["pdf"].get("element") or m["pdf"].get("section", "")),
                    "subsection": m["pdf"].get("subsection", ""),
                    "field": m["pdf"].get("field", ""),
                    "pdfValue": m["pdf_value"],
                    "excelValue": m["excel_value"],
                    "excelRow": m["excel"].get("row"),
                }
                for m in mismatches
            ],
            "matched": [
                {
                    "section": (m["pdf"].get("element") or m["pdf"].get("section", "")),
                    "subsection": m["pdf"].get("subsection", ""),
                    "field": m["pdf"].get("field", ""),
                    "pdfValue": m["pdf"].get("permissions_str", ""),
                    "excelValue": m["excel"].get("value", ""),
                    "excelRow": m["excel"].get("row"),
                }
                for m in matched
            ],
        }

    finally:
        if pdf_tmp and os.path.exists(pdf_tmp.name):
            os.unlink(pdf_tmp.name)
        if excel_tmp and os.path.exists(excel_tmp.name):
            os.unlink(excel_tmp.name)


@app.post("/api/modify-excel")
async def modify_excel(
    excel: UploadFile = File(...),
    rows: str = Form(...),
    role_column: int = Form(...),
):
    """
    Accept an Excel workbook, a comma-separated list of row numbers, and a
    role column index.  Set each specified cell to 'None' and return the
    modified workbook as a download.
    """

    row_nums = [int(r) for r in rows.split(",") if r.strip().isdigit()]
    if not row_nums:
        raise HTTPException(status_code=400, detail="No valid row numbers provided")

    excel_tmp = None
    try:
        excel_tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        excel_tmp.write(await excel.read())
        excel_tmp.close()

        wb = openpyxl.load_workbook(excel_tmp.name)
        EXCEL_SHEET = "ROLE ACCESS (WHAT)"
        if EXCEL_SHEET not in wb.sheetnames:
            raise HTTPException(status_code=422, detail=f"Sheet '{EXCEL_SHEET}' not found")

        ws = wb[EXCEL_SHEET]
        for row_num in row_nums:
            ws.cell(row=row_num, column=role_column).value = "None"

        # Save to bytes
        out = io.BytesIO()
        wb.save(out)
        wb.close()
        out.seek(0)

        download_name = os.path.splitext(excel.filename or "workbook")[0] + "_modified.xlsx"

        return StreamingResponse(
            out,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{download_name}"'},
        )
    finally:
        if excel_tmp and os.path.exists(excel_tmp.name):
            os.unlink(excel_tmp.name)


@app.post("/api/update-excel-from-pdf")
async def update_excel_from_pdf(
    pdf: UploadFile = File(...),
    excel: UploadFile = File(...),
):
    """
    Accept a PDF + Excel, find Excel cells where the field exists but value is
    None, update those cells with the PDF permission values, and return the
    modified Excel file for download.
    """
    pdf_tmp = None
    excel_tmp = None
    try:
        pdf_tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        pdf_tmp.write(await pdf.read())
        pdf_tmp.close()

        excel_tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        excel_tmp.write(await excel.read())
        excel_tmp.close()

        # Extract PDF permissions
        try:
            pdf_entries = extract_pdf_permissions(pdf_tmp.name)
        except Exception as exc:
            raise HTTPException(status_code=422, detail=f"PDF parsing error: {exc}") from exc

        if not pdf_entries:
            raise HTTPException(status_code=422, detail="No permission entries found in PDF.")

        # Detect role name from PDF filename
        pdf_basename = os.path.splitext(pdf.filename or "")[0]
        role_name = None
        import re as _re
        m = _re.search(r'View Role for\s+(.+)', pdf_basename, _re.IGNORECASE)
        if m:
            role_name = m.group(1).strip()
        if not role_name:
            role_name = pdf_basename.strip()

        EXCEL_SHEET = "ROLE ACCESS (WHAT)"

        # Open Excel and find role column
        try:
            wb = openpyxl.load_workbook(excel_tmp.name)
        except Exception as exc:
            raise HTTPException(status_code=422, detail=f"Excel open error: {exc}") from exc

        if EXCEL_SHEET not in wb.sheetnames:
            wb.close()
            raise HTTPException(
                status_code=422,
                detail=f"Sheet '{EXCEL_SHEET}' not found in Excel workbook.",
            )

        ws = wb[EXCEL_SHEET]
        role_col = None
        role_name_norm = role_name.lower().strip()
        for header_row in range(1, 4):
            for col in range(1, ws.max_column + 1):
                cell_val = ws.cell(row=header_row, column=col).value
                if cell_val and role_name_norm in str(cell_val).lower().strip():
                    role_col = col
                    break
            if role_col:
                break

        if role_col is None:
            wb.close()
            raise HTTPException(
                status_code=422,
                detail=f"Role '{role_name}' not found in header rows of sheet '{EXCEL_SHEET}'.",
            )

        # Extract Excel permissions and compare
        excel_entries = extract_excel_permissions(excel_tmp.name, EXCEL_SHEET, role_col)
        missing_in_excel, _, _, _ = compare_pdf_vs_excel(pdf_entries, excel_entries)

        has_field_none = [e for e in missing_in_excel if e.get("_excel_has_field")]

        if not has_field_none:
            wb.close()
            raise HTTPException(
                status_code=422,
                detail="No 'Excel field = None' entries found to update.",
            )

        # Update the Excel cells
        updated_count = 0
        for entry in has_field_none:
            row = entry.get("_excel_row")
            pdf_value = entry.get("permissions_str", "")
            if row and pdf_value:
                ws.cell(row=row, column=role_col).value = pdf_value
                updated_count += 1

        # Save modified workbook
        out_tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        out_tmp.close()
        wb.save(out_tmp.name)
        wb.close()

        with open(out_tmp.name, "rb") as fh:
            xlsx_bytes = fh.read()
        os.unlink(out_tmp.name)

        excel_basename = os.path.splitext(excel.filename or "workbook")[0]
        download_name = f"{excel_basename}_updated.xlsx"

        return StreamingResponse(
            io.BytesIO(xlsx_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{download_name}"'},
        )

    finally:
        if pdf_tmp and os.path.exists(pdf_tmp.name):
            os.unlink(pdf_tmp.name)
        if excel_tmp and os.path.exists(excel_tmp.name):
            os.unlink(excel_tmp.name)


@app.post("/api/pdf-to-excel")
async def pdf_to_excel(pdf: UploadFile = File(...)):
    """
    Accept a PDF role export, extract all permissions, and return a
    downloadable Excel (.xlsx) file with the permission data.
    """
    pdf_tmp = None
    xlsx_tmp = None
    try:
        pdf_tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        pdf_tmp.write(await pdf.read())
        pdf_tmp.close()

        try:
            pdf_entries = extract_pdf_permissions(pdf_tmp.name)
        except Exception as exc:
            raise HTTPException(status_code=422, detail=f"PDF parsing error: {exc}") from exc

        if not pdf_entries:
            raise HTTPException(status_code=422, detail="No permission entries found in PDF.")

        # Generate Excel output
        xlsx_tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        xlsx_tmp.close()

        try:
            generate_pdf_excel(pdf_entries, xlsx_tmp.name)
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"Excel generation error: {exc}") from exc

        # Read into memory so we can clean up the temp file before returning
        with open(xlsx_tmp.name, "rb") as fh:
            xlsx_bytes = fh.read()

        # Build a safe download filename from the PDF name
        pdf_basename = os.path.splitext(pdf.filename or "permissions")[0]
        download_name = pdf_basename + "_permissions.xlsx"

        return StreamingResponse(
            io.BytesIO(xlsx_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{download_name}"'},
        )

    finally:
        if pdf_tmp and os.path.exists(pdf_tmp.name):
            os.unlink(pdf_tmp.name)
        if xlsx_tmp and os.path.exists(xlsx_tmp.name):
            os.unlink(xlsx_tmp.name)
