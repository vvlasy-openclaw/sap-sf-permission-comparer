"""
==============================================================================
PDF vs Excel Permission Comparator
==============================================================================
Compares permissions extracted from the PDF (system export) against the
Excel workbook's "ROLE ACCESS (WHAT)" sheet for the ALL_MGR_GL_Manager role.

Produces a report of:
  1. Permissions in PDF but MISSING from Excel
  2. Permissions in Excel but MISSING from PDF
  3. Permission VALUE mismatches (same field, different access level)

Requirements:
    pip install PyMuPDF openpyxl

Usage:
    python compare_pdf_vs_excel.py
==============================================================================
"""

import os
import sys
import re
from datetime import datetime

import openpyxl

# Reuse PDF parsing from comparer.py
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from comparer import extract_pdf_structured

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# --------------------------------------------------------------------------
# CONFIGURATION
# --------------------------------------------------------------------------
PDF_PATH = os.path.join(SCRIPT_DIR, "View Role for ALL_MGR_GL_Manager NEW.pdf")
EXCEL_PATH = os.path.join(
    SCRIPT_DIR,
    "BBraun_Role Based Permission Workbook (version 1) (1).xlsx",
)
EXCEL_SHEET = "ROLE ACCESS (WHAT)"
ROLE_COLUMN = 12  # Column L = ALL_MGR_GL_Manager

# Sections to skip (metadata, not permissions)
SKIP_SECTIONS = {
    "__HEADER__", "User Type", "Role Name", "Role Description",
    "RBP-Only", "Last Modified By", "Last Modified Date",
}


# --------------------------------------------------------------------------
# PDF EXTRACTION
# --------------------------------------------------------------------------

def extract_pdf_permissions(pdf_path):
    """
    Extract all permissions from the PDF using PyMuPDF's layout-based
    extraction and return a list of dicts:
        [{section, element, subsection, field, permissions_set, permissions_str}, ...]
    """
    raw_entries = extract_pdf_structured(pdf_path)
    entries = []

    for raw in raw_entries:
        section = raw["section"]
        element = raw["element"]
        subsection = raw["subsection"]
        field = raw["field"]
        perms_str = raw["permissions_str"]

        # Skip metadata sections
        if section in SKIP_SECTIONS:
            continue

        # --- Handle Field-Level Permissions blocks ---
        # These have raw override text like:
        # "Field-Level Permissions Higher Level Position (parentPosition)=Read-Only | ..."
        if "Field-Level Permissions" in field:
            # Parse the override block
            override_text = field
            if perms_str:
                override_text = override_text + " " + perms_str
            # Strip the "Field-Level Permissions " prefix
            override_text = re.sub(r'^Field-Level Permissions\s+', '', override_text)
            # Normalize wrapping artifacts
            override_text = re.sub(r'Read-\s+Only', 'Read-Only', override_text)
            override_text = re.sub(r'No\s+Access', 'No Access', override_text)
            # Split on | to get individual overrides
            parts = re.split(r'\s*\|\s*', override_text)
            overrides = []
            for part in parts:
                part = part.strip()
                if not part:
                    continue
                match = re.match(r'^(.+?)\s*\([^)]*\)\s*=\s*(.+)$', part)
                if match:
                    fname = match.group(1).strip()
                    value = match.group(2).strip().replace("Read-Only", "Read Only")
                    overrides.append((fname, value))
                else:
                    match2 = re.match(r'^(.+?)\s*=\s*(.+)$', part)
                    if match2:
                        fname = match2.group(1).strip()
                        value = match2.group(2).strip().replace("Read-Only", "Read Only")
                        overrides.append((fname, value))
            if overrides:
                # Add "Field Level Overrides = Yes" entry
                entries.append({
                    "section": section,
                    "element": element,
                    "subsection": subsection,
                    "field": "Field Level Overrides",
                    "permissions_set": {"Enabled"},
                    "permissions_str": "Yes",
                })
                for override_field, override_value in overrides:
                    entries.append({
                        "section": section,
                        "element": element,
                        "subsection": subsection,
                        "field": "Field Level Overrides - " + override_field,
                        "permissions_set": {override_value},
                        "permissions_str": override_value,
                    })
            continue

        # --- Handle Object-Level Permissions ---
        # In the PDF, entries like "Department > Object-Level Permissions = View Current"
        # map in Excel to field="Department", value="View Current".
        # The "Object-Level Permissions" label is just structural — the real field
        # name is the parent (subsection).
        if field == "Object-Level Permissions" and subsection:
            field = subsection

        # --- Transform †(...) suffix into Excel "- All" / "- Others" format ---
        # "Goal Plan Permissions †(All)" → "Goal Plan Permissions - All"
        # "Goal Plan Permissions †(Learning Activity, ...)" → "Goal Plan Permissions - Others"
        dagger_paren = re.search(r'\s*†\((.+)\)\s*$', field)
        if dagger_paren:
            paren_content = dagger_paren.group(1).strip()
            base_name = field[:dagger_paren.start()]
            if paren_content.lower() == "all":
                field = base_name + " - All"
            else:
                field = base_name + " - Others"
        else:
            # --- Non-dagger parenthesized items (semicolon-separated) ---
            # "Form Filters(Template; Current Step, ...)" → "Form Filters - Other"
            bare_paren = re.search(r'(\w)\((.+)\)\s*$', field)
            if bare_paren and ';' in bare_paren.group(2):
                paren_content = bare_paren.group(2).strip()
                # Keep the last char of the base name that was matched by (\w)
                base_name = field[:bare_paren.start() + 1]
                if paren_content.lower() == "all":
                    field = base_name + " - All"
                else:
                    field = base_name + " - Other"

        # --- Handle Talent Search "Searchable" suffix ---
        is_searchable_field = (
            field.endswith(" Searchable")
            and (perms_str == "Enabled" or not perms_str)
        )
        if is_searchable_field:
            field = field[:-len(" Searchable")]
            perms_str = "Searchable"

        # --- Parse the permission string into a set ---
        if perms_str == "Enabled":
            rights = {"Enabled"}
            perm_str = normalize_pdf_permissions(rights, field)
        elif perms_str == "Searchable":
            rights = {"Searchable"}
            perm_str = "Searchable"
        elif perms_str:
            # Parse "View Current | View History | Edit/Insert" into a set
            parts = re.split(r'\s*\|\s*', perms_str)
            rights = set()
            for p in parts:
                p = p.strip().rstrip(" †⁜★")
                if p:
                    rights.add(p)
            perm_str = normalize_pdf_permissions(rights, field)
        else:
            rights = {"Enabled"}
            perm_str = "Yes"

        # Filter artifacts
        if _is_pdf_artifact(field, rights):
            continue

        # For Object-Level Permissions entries, also look for the next field-level
        # override entry (handled above via "Field-Level Permissions" block)

        entries.append({
            "section": section,
            "element": element,
            "subsection": subsection,
            "field": field,
            "permissions_set": rights,
            "permissions_str": perm_str,
        })

    return entries




def _is_pdf_artifact(field, rights):
    """
    Detect PDF parsing artifacts from multi-line text wrapping.
    These are fragment lines that got parsed as separate entries.
    """
    # Lines starting with "(" are ID-code continuations from the previous line
    # e.g., "(phased retirement-presence) (JC_DEU_ATZ2)"
    if field.startswith("(") and rights == {"Enabled"}:
        return True

    # Lines containing "=Read-Only" or "=No Access" are field-level override
    # fragments from the Position section that got split across lines
    if "=" in field and ("Read-Only" in field or "No Access" in field):
        return True

    # Empty permissions (unparseable fragments)
    if not rights:
        return True

    return False


def normalize_pdf_permissions(rights_set, field_name=""):
    """
    Convert a set of permissions like {'View Current', 'View History', 'Edit/Insert'}
    into a normalized string like 'View Current + View History + Edit/Insert'
    that matches the Excel format.
    """
    if rights_set == {"Enabled"}:
        # Fields ending in "Searchable" in the Reports Permission section
        # map to "Searchable" in Excel, not "Yes"
        if field_name.lower().endswith("searchable"):
            return "Searchable"
        return "Yes"

    # Define ordering for consistency
    order = [
        "View", "View Current", "View History",
        "Edit", "Edit/Insert", "Insert",
        "Create", "Correct", "Delete",
        "Import", "Export", "Import/Export",
        "Approve", "Admin", "None",
        "Yes", "No", "Searchable", "Read Only",
    ]

    sorted_perms = sorted(rights_set, key=lambda x: order.index(x) if x in order else 999)
    return " + ".join(sorted_perms)


# --------------------------------------------------------------------------
# EXCEL EXTRACTION
# --------------------------------------------------------------------------
def extract_excel_permissions(excel_path, sheet_name, role_col):
    """
    Read the Excel ROLE ACCESS (WHAT) sheet and return a list of dicts:
        [{category, element, grouping, permission, value, row}, ...]
    Only includes rows where the role column has an actual permission value.
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name]
    entries = []

    # Track "inherited" category/element values (they can be blank in some rows)
    last_category = ""
    last_element = ""

    for row_idx in range(4, ws.max_row + 1):  # Data starts at row 4
        cat = ws.cell(row=row_idx, column=2).value
        elem = ws.cell(row=row_idx, column=3).value
        grp = ws.cell(row=row_idx, column=4).value
        perm = ws.cell(row=row_idx, column=5).value
        val = ws.cell(row=row_idx, column=role_col).value

        # Carry forward category/element from previous rows
        if cat and str(cat).strip():
            last_category = str(cat).strip()
            # Section header rows have cat but no elem/grp/perm — use the
            # category as the element for subsequent rows that lack an element.
            if not (elem and str(elem).strip()) and not (perm and str(perm).strip()):
                last_element = last_category.title()
        if elem and str(elem).strip():
            last_element = str(elem).strip()

        # Skip rows without a permission name
        if not perm or not str(perm).strip():
            continue

        perm_str = str(perm).strip()
        grp_str = str(grp).strip() if grp else ""

        # Get the permission value for this role
        val_str = str(val).strip() if val else "None"

        # Skip metadata rows
        if val_str in ("Alex (done)", "Alex (Zuzka working)"):
            continue

        entries.append({
            "category": last_category,
            "element": last_element,
            "grouping": grp_str,
            "permission": perm_str,
            "value": val_str,
            "row": row_idx,
        })

    wb.close()
    return entries


# --------------------------------------------------------------------------
# MATCHING LOGIC
# --------------------------------------------------------------------------

def build_excel_lookup(excel_entries):
    """
    Build lookup dictionaries from Excel entries for matching against PDF.
    Returns:
        by_field: {normalized_field_name: [entries]}
        by_group_field: {(grouping, field): [entries]}
        by_group_field_contains: flat list for substring matching
    """
    by_field = {}
    by_group_field = {}

    for entry in excel_entries:
        field_key = _normalize_for_match(entry["permission"])
        if field_key not in by_field:
            by_field[field_key] = []
        by_field[field_key].append(entry)

        grp_key = (_normalize_for_match(entry["grouping"]), field_key)
        if grp_key not in by_group_field:
            by_group_field[grp_key] = []
        by_group_field[grp_key].append(entry)

    return by_field, by_group_field


def _normalize_for_match(s):
    """Normalize a string for fuzzy matching: lowercase, strip markers, collapse spaces."""
    s = s.lower().strip()
    s = re.sub(r'[†⁜★]', '', s)
    # Normalize curly/smart quotes to straight quotes
    s = s.replace('\u201c', '"').replace('\u201d', '"')
    s = s.replace('\u2018', "'").replace('\u2019', "'")
    s = re.sub(r'\s+', ' ', s)
    return s.strip()


def find_excel_match(pdf_entry, by_field, by_group_field, excel_entries):
    """
    Try to find the best matching Excel entry for a PDF permission entry.
    Returns the matched Excel entry or None.
    """
    field = _normalize_for_match(pdf_entry["field"])
    subsection = _normalize_for_match(pdf_entry["subsection"])
    section = pdf_entry["section"].lower()
    element = _normalize_for_match(pdf_entry.get("element", ""))

    def _pick_best(candidates):
        """From a list of Excel candidates, pick the best match using element context."""
        if len(candidates) == 1:
            return candidates[0]
        # Try to narrow by element context (bold header in PDF = Element in Excel)
        if element:
            for c in candidates:
                if _normalize_for_match(c["element"]) == element:
                    return c
        # Try to narrow by section/element context
        for c in candidates:
            if c["element"].lower() in section or section in c["element"].lower():
                return c
        # Try to narrow by grouping matching subsection
        for c in candidates:
            if subsection and _normalize_for_match(c["grouping"]) == subsection:
                return c
        # Return first match as fallback
        return candidates[0]

    # Try exact match on (grouping, field) first
    key = (subsection, field)
    if key in by_group_field:
        return _pick_best(by_group_field[key])

    # Try field name only
    if field in by_field:
        return _pick_best(by_field[field])

    # Parent-child suffix matching: PDF has "User Search †" as a child of
    # "Company Info Access †", but Excel has "Company Info Access - User Search".
    # Try matching "X - field" patterns in the same grouping.
    suffix_pattern = " - " + field
    for entry in excel_entries:
        norm_eperm = _normalize_for_match(entry["permission"])
        if norm_eperm.endswith(suffix_pattern):
            norm_grp = _normalize_for_match(entry["grouping"])
            if (not subsection
                    or norm_grp == subsection
                    or norm_grp == element
                    or _normalize_for_match(entry["element"]) == element):
                return entry

    # For "Object-Level Permissions" in MDF/Miscellaneous sections, match by subsection=grouping
    if field == "object-level permissions" and subsection:
        for entry in excel_entries:
            if (_normalize_for_match(entry["permission"]) == "object-level permissions"
                and _normalize_for_match(entry["grouping"]).startswith(_normalize_for_match(subsection))):
                return entry

    # Substring matching: PDF field might be truncated (multi-line wrap)
    # e.g., PDF "AUT - Vacation contingent Start" should match Excel "AUT - Vacation contingent Start Date"
    if len(field) > 10:
        norm_field = _normalize_for_match(field)
        best = None
        best_len = 0
        for entry in excel_entries:
            norm_eperm = _normalize_for_match(entry["permission"])
            # Check if the PDF field is a prefix of an Excel field in the same grouping
            if subsection and _normalize_for_match(entry["grouping"]) != _normalize_for_match(subsection):
                continue
            if norm_eperm.startswith(norm_field) and len(norm_eperm) > best_len:
                best = entry
                best_len = len(norm_eperm)
            elif norm_field.startswith(norm_eperm) and len(norm_eperm) > best_len:
                best = entry
                best_len = len(norm_eperm)
        if best:
            return best

    # For Event Reasons: the Excel has "Name (ID_CODE)" as one field,
    # but the PDF splits into "Name" on one line and "(ID_CODE)" as a separate entry.
    # Try matching by checking if any Excel field contains the PDF field name.
    if subsection and "event reason" in subsection.lower():
        norm_field = _normalize_for_match(field)
        for entry in excel_entries:
            if _normalize_for_match(entry["grouping"]) != _normalize_for_match(subsection):
                continue
            norm_eperm = _normalize_for_match(entry["permission"])
            if norm_field in norm_eperm or norm_eperm in norm_field:
                return entry

    return None


def normalize_excel_value(val):
    """Normalize Excel permission value for comparison."""
    if not val or val == "None" or val == "None (not in system)":
        return "None"
    return val.strip()


def compare_pdf_vs_excel(pdf_entries, excel_entries):
    """
    Compare PDF permissions against Excel and return:
        missing_in_excel: PDF entries not found in Excel
        missing_in_pdf: Excel entries not found in PDF
        mismatches: entries found in both but with different permission values
        matched: entries that match perfectly
    """
    by_field, by_group_field = build_excel_lookup(excel_entries)

    missing_in_excel = []
    mismatches = []
    matched = []
    pdf_matched_excel_rows = set()

    for pdf_entry in pdf_entries:
        excel_match = find_excel_match(pdf_entry, by_field, by_group_field, excel_entries)

        if excel_match is None:
            missing_in_excel.append(pdf_entry)
            continue

        pdf_matched_excel_rows.add(excel_match["row"])

        # Compare permission values
        pdf_val = pdf_entry["permissions_str"]
        excel_val = normalize_excel_value(excel_match["value"])

        if excel_val == "None":
            # Excel has the field but with "None" permission — treat as missing
            missing_in_excel.append({
                **pdf_entry,
                "_excel_row": excel_match["row"],
                "_excel_has_field": True,
                "_excel_value": excel_val,
            })
        elif not permissions_match(pdf_val, excel_val):
            mismatches.append({
                "pdf": pdf_entry,
                "excel": excel_match,
                "pdf_value": pdf_val,
                "excel_value": excel_val,
            })
        else:
            matched.append({
                "pdf": pdf_entry,
                "excel": excel_match,
            })

    # Find Excel entries not matched to any PDF entry (with actual permissions)
    missing_in_pdf = []
    for entry in excel_entries:
        if entry["row"] not in pdf_matched_excel_rows:
            if entry["value"] not in ("None", "None (not in system)", "No", "No Access"):
                missing_in_pdf.append(entry)

    return missing_in_excel, missing_in_pdf, mismatches, matched


def permissions_match(pdf_val, excel_val):
    """
    Check if PDF and Excel permission values are equivalent.
    Handles format differences:
        PDF: "View | Edit" or "View Current + View History"
        Excel: "View + Edit" or "View Current + View History + Edit/Insert"
    """
    # Normalize both to a comparable set
    pdf_set = parse_perm_string(pdf_val)
    excel_set = parse_perm_string(excel_val)
    return pdf_set == excel_set


def parse_perm_string(val):
    """Parse a permission string into a normalized set of individual permissions."""
    if not val or val == "None":
        return set()

    # Handle "Yes" / "Enabled" as equivalent
    if val.strip().lower() in ("yes", "enabled"):
        return {"Yes"}

    # Split by + or | (both formats)
    parts = re.split(r'\s*[+|]\s*', val)
    result = set()
    for p in parts:
        p = p.strip().rstrip(" ★†⁜")
        if p:
            result.add(p)
    return result


# --------------------------------------------------------------------------
# REPORT GENERATION
# --------------------------------------------------------------------------

def generate_report(missing_in_excel, missing_in_pdf, mismatches, matched,
                    total_pdf=0, total_excel=0, total_excel_active=0):
    """Generate a formatted console and text file report."""
    sep = "=" * 90
    sub_sep = "-" * 90
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    lines = [
        sep,
        "  PDF vs EXCEL PERMISSION COMPARISON REPORT",
        "  Role: ALL_MGR_GL_Manager",
        sep,
        f"  Generated: {timestamp}",
        f"  PDF:   {os.path.basename(PDF_PATH)}",
        f"  Excel: {os.path.basename(EXCEL_PATH)}",
        sep,
        "",
        "  SUMMARY",
        sub_sep,
        f"  Total PDF permission entries:     {total_pdf}",
        f"  Total Excel permission entries:   {total_excel}",
        f"    of which with active values:    {total_excel_active}",
        "",
        f"  Matched (identical):              {len(matched)}",
        f"  In PDF, missing/None in Excel:    {len(missing_in_excel)}",
        f"  In Excel, not found in PDF:       {len(missing_in_pdf)}",
        f"  Permission value mismatches:      {len(mismatches)}",
        "",
    ]

    def _fmt_entry(entry):
        """Format a PDF entry for display, using element as section when available."""
        section = entry.get("element") or entry["section"]
        subsection = entry['subsection']
        # Don't repeat the subsection if it's the same as the section/element
        if subsection and subsection != section:
            return f"[{section}] {subsection} > {entry['field']}"
        return f"[{section}] {entry['field']}"

    # --- Missing in Excel ---
    if missing_in_excel:
        lines.append(f"\n  IN PDF BUT MISSING/NONE IN EXCEL ({len(missing_in_excel)} items)")
        lines.append(sub_sep)

        # Separate into truly missing vs "field exists but value is None"
        truly_missing = [e for e in missing_in_excel if not e.get("_excel_has_field")]
        has_field_none = [e for e in missing_in_excel if e.get("_excel_has_field")]

        if truly_missing:
            lines.append(f"\n  --- Field NOT FOUND in Excel ({len(truly_missing)}) ---")
            for i, entry in enumerate(truly_missing, 1):
                lines.append(f"    {i}. {_fmt_entry(entry)}")
                lines.append(f"       PDF permissions: {entry['permissions_str']}")

        if has_field_none:
            lines.append(f"\n  --- Field EXISTS in Excel but value is None ({len(has_field_none)}) ---")
            lines.append(f"       (These need their permission value updated in the Excel)")
            for i, entry in enumerate(has_field_none, 1):
                lines.append(f"    {i}. {_fmt_entry(entry)}")
                lines.append(f"       PDF permissions: {entry['permissions_str']}")
                lines.append(f"       Excel row: {entry.get('_excel_row', '?')}")

    # --- Missing in PDF ---
    if missing_in_pdf:
        lines.append(f"\n  IN EXCEL BUT NOT FOUND IN PDF ({len(missing_in_pdf)} items)")
        lines.append(sub_sep)
        for i, entry in enumerate(missing_in_pdf, 1):
            lines.append(f"    {i}. [{entry['element']}] {entry['grouping']} > {entry['permission']}")
            lines.append(f"       Excel value: {entry['value']}  (row {entry['row']})")

    # --- Mismatches ---
    if mismatches:
        lines.append(f"\n  PERMISSION VALUE MISMATCHES ({len(mismatches)} items)")
        lines.append(sub_sep)
        for i, m in enumerate(mismatches, 1):
            pdf_e = m["pdf"]
            excel_e = m["excel"]
            lines.append(f"    {i}. {_fmt_entry(pdf_e)}")
            lines.append(f"       PDF:   {m['pdf_value']}")
            lines.append(f"       Excel: {m['excel_value']}  (row {excel_e['row']})")

    lines.append("")
    lines.append(sep)
    lines.append("  END OF REPORT")
    lines.append(sep)

    return "\n".join(lines)


def _html_escape(text):
    """Escape HTML special characters."""
    return (str(text).replace("&", "&amp;").replace("<", "&lt;")
                     .replace(">", "&gt;").replace('"', "&quot;"))


def _fmt_entry_parts(entry):
    """Return (section, subsection, field) for display."""
    section = entry.get("element") or entry["section"]
    subsection = entry.get("subsection", "")
    if subsection and subsection != section:
        return section, subsection, entry["field"]
    return section, "", entry["field"]


def generate_html_report(missing_in_excel, missing_in_pdf, mismatches, matched,
                         total_pdf=0, total_excel=0, total_excel_active=0):
    """Generate a colour-coded HTML report matching the T3/PROD style."""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    pdf_name = _html_escape(os.path.basename(PDF_PATH))
    excel_name = _html_escape(os.path.basename(EXCEL_PATH))

    truly_missing = [e for e in missing_in_excel if not e.get("_excel_has_field")]
    has_field_none = [e for e in missing_in_excel if e.get("_excel_has_field")]

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>PDF vs Excel Permission Comparison — ALL_MGR_GL_Manager</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: 'Segoe UI', system-ui, -apple-system, sans-serif; background: #f5f6fa; color: #2c3e50; line-height: 1.5; }}
  .container {{ max-width: 1200px; margin: 0 auto; padding: 24px; }}
  header {{ background: linear-gradient(135deg, #2c3e50, #8e44ad); color: white; padding: 32px; border-radius: 12px; margin-bottom: 24px; }}
  header h1 {{ font-size: 1.6em; margin-bottom: 8px; }}
  header .meta {{ opacity: 0.85; font-size: 0.9em; }}
  header .meta span {{ display: inline-block; margin-right: 24px; }}
  .summary {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(160px, 1fr)); gap: 16px; margin-bottom: 28px; }}
  .card {{ background: white; border-radius: 10px; padding: 20px; box-shadow: 0 2px 8px rgba(0,0,0,0.06); text-align: center; border-top: 4px solid #bdc3c7; }}
  .card .num {{ font-size: 2.2em; font-weight: 700; }}
  .card .label {{ font-size: 0.85em; color: #7f8c8d; margin-top: 4px; }}
  .section {{ background: white; border-radius: 10px; box-shadow: 0 2px 8px rgba(0,0,0,0.06); margin-bottom: 20px; overflow: hidden; }}
  .section-header {{ padding: 16px 20px; font-weight: 600; font-size: 1.05em; cursor: pointer; display: flex; justify-content: space-between; align-items: center; color: white; }}
  .section-header .badge {{ background: rgba(255,255,255,0.3); padding: 2px 10px; border-radius: 12px; font-size: 0.85em; }}
  .section-body {{ padding: 0; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 0.9em; }}
  th {{ background: #f8f9fa; padding: 10px 16px; text-align: left; font-weight: 600; color: #555; border-bottom: 2px solid #eee; position: sticky; top: 0; }}
  td {{ padding: 10px 16px; border-bottom: 1px solid #f0f0f0; vertical-align: top; }}
  tr:hover td {{ background: #fafbfc; }}
  .tag {{ display: inline-block; padding: 2px 8px; border-radius: 4px; font-size: 0.8em; font-weight: 500; }}
  .tag-pdf {{ background: #eaf2f8; color: #2471a3; }}
  .tag-excel {{ background: #eafaf1; color: #1e8449; }}
  .tag-mismatch {{ background: #fef5e7; color: #d35400; }}
  .tag-none {{ background: #f2f3f4; color: #95a5a6; }}
  .group-row td {{ background: #34495e; color: white; font-weight: 600; font-size: 0.9em; padding: 8px 16px; border-bottom: none; }}
  .group-row:hover td {{ background: #34495e; }}
  .filter-bar {{ background: white; border-radius: 10px; box-shadow: 0 2px 8px rgba(0,0,0,0.06); padding: 16px 20px; margin-bottom: 20px; display: flex; gap: 12px; align-items: center; flex-wrap: wrap; }}
  .filter-bar label {{ font-weight: 600; font-size: 0.9em; color: #555; white-space: nowrap; }}
  .filter-bar input {{ flex: 1; min-width: 200px; padding: 8px 12px; border: 1px solid #ddd; border-radius: 6px; font-size: 0.9em; }}
  .filter-bar input:focus {{ outline: none; border-color: #8e44ad; box-shadow: 0 0 0 2px rgba(142,68,173,0.2); }}
  .filter-btn {{ padding: 6px 14px; border: 1px solid #ddd; border-radius: 6px; background: #f8f9fa; cursor: pointer; font-size: 0.82em; color: #555; transition: all 0.15s; }}
  .filter-btn:hover {{ background: #e8e8e8; }}
  .filter-btn.active {{ background: #8e44ad; color: white; border-color: #8e44ad; }}
  .filter-info {{ font-size: 0.82em; color: #95a5a6; margin-left: auto; white-space: nowrap; }}
  .hidden-row {{ display: none; }}
  .collapsible {{ cursor: pointer; user-select: none; }}
  .collapsible .arrow {{ transition: transform 0.2s; display: inline-block; margin-right: 6px; }}
  .collapsed .arrow {{ transform: rotate(-90deg); }}
  .hidden-body {{ display: none; }}
</style>
</head>
<body>
<div class="container">
  <header>
    <h1>PDF vs Excel Permission Comparison</h1>
    <div class="meta">
      <span>Role: ALL_MGR_GL_Manager</span>
      <span>Generated: {timestamp}</span>
    </div>
    <div class="meta" style="margin-top: 6px;">
      <span>PDF: {pdf_name}</span>
      <span>Excel: {excel_name}</span>
    </div>
  </header>

  <div class="summary">
    <div class="card" style="border-top-color: #27ae60;">
      <div class="num">{len(matched)}</div>
      <div class="label">Matched</div>
    </div>
    <div class="card" style="border-top-color: #e67e22;">
      <div class="num">{len(mismatches)}</div>
      <div class="label">Mismatches</div>
    </div>
    <div class="card" style="border-top-color: #e74c3c;">
      <div class="num">{len(truly_missing)}</div>
      <div class="label">Not in Excel</div>
    </div>
    <div class="card" style="border-top-color: #9b59b6;">
      <div class="num">{len(has_field_none)}</div>
      <div class="label">Excel = None</div>
    </div>
    <div class="card" style="border-top-color: #3498db;">
      <div class="num">{len(missing_in_pdf)}</div>
      <div class="label">Not in PDF</div>
    </div>
    <div class="card" style="border-top-color: #95a5a6;">
      <div class="num">{total_pdf}</div>
      <div class="label">PDF Entries</div>
    </div>
  </div>

  <div class="filter-bar">
    <label>Filter:</label>
    <input type="text" id="filterInput" placeholder="Type to search/filter rows...">
    <button class="filter-btn" onclick="clearFilter()">Clear</button>
    <span class="filter-info" id="filterInfo"></span>
  </div>
"""

    # --- Helper: render rows with section group separators ---
    def _render_pdf_rows(items, get_parts_fn, cols_fn, col_count):
        """Render table rows with section group header rows when section changes."""
        rows_html = ""
        last_sec = None
        for i, item in enumerate(items, 1):
            sec, sub, fld = get_parts_fn(item)
            if sec != last_sec:
                rows_html += f'        <tr class="group-row"><td colspan="{col_count}">{_html_escape(sec)}</td></tr>\n'
                last_sec = sec
            rows_html += cols_fn(i, item, sub, fld)
        return rows_html

    # --- Matched section ---
    if matched:
        html += f"""
  <div class="section">
    <div class="section-header collapsible collapsed" style="background: #27ae60;" onclick="toggleSection(this)">
      <span><span class="arrow">&#9660;</span>Matched Permissions</span>
      <span class="badge">{len(matched)}</span>
    </div>
    <div class="section-body hidden-body">
      <table>
        <tr><th>#</th><th>Subsection</th><th>Permission</th><th>Value</th><th>Excel Row</th></tr>
"""
        def _matched_cols(i, m, sub, fld):
            val = _html_escape(m["pdf"]["permissions_str"])
            row = m["excel"]["row"]
            return f'        <tr><td>{i}</td><td>{_html_escape(sub)}</td><td>{_html_escape(fld)}</td><td><span class="tag tag-pdf">{val}</span></td><td>{row}</td></tr>\n'
        html += _render_pdf_rows(matched, lambda m: _fmt_entry_parts(m["pdf"]), _matched_cols, 5)
        html += """      </table>
    </div>
  </div>
"""

    # --- Mismatches section ---
    if mismatches:
        html += f"""
  <div class="section">
    <div class="section-header collapsible" style="background: #e67e22;" onclick="toggleSection(this)">
      <span><span class="arrow">&#9660;</span>Permission Value Mismatches</span>
      <span class="badge">{len(mismatches)}</span>
    </div>
    <div class="section-body">
      <table>
        <tr><th>#</th><th>Subsection</th><th>Permission</th><th>PDF Value</th><th>Excel Value</th><th>Excel Row</th></tr>
"""
        def _mismatch_cols(i, m, sub, fld):
            pdf_val = _html_escape(m["pdf_value"])
            excel_val = _html_escape(m["excel_value"])
            row = m["excel"]["row"]
            return f'        <tr><td>{i}</td><td>{_html_escape(sub)}</td><td>{_html_escape(fld)}</td><td><span class="tag tag-pdf">{pdf_val}</span></td><td><span class="tag tag-excel">{excel_val}</span></td><td>{row}</td></tr>\n'
        html += _render_pdf_rows(mismatches, lambda m: _fmt_entry_parts(m["pdf"]), _mismatch_cols, 6)
        html += """      </table>
    </div>
  </div>
"""

    # --- Not found in Excel ---
    if truly_missing:
        html += f"""
  <div class="section">
    <div class="section-header collapsible" style="background: #e74c3c;" onclick="toggleSection(this)">
      <span><span class="arrow">&#9660;</span>In PDF but NOT FOUND in Excel</span>
      <span class="badge">{len(truly_missing)}</span>
    </div>
    <div class="section-body">
      <table>
        <tr><th>#</th><th>Subsection</th><th>Permission</th><th>PDF Value</th></tr>
"""
        def _missing_excel_cols(i, entry, sub, fld):
            val = _html_escape(entry["permissions_str"])
            return f'        <tr><td>{i}</td><td>{_html_escape(sub)}</td><td>{_html_escape(fld)}</td><td><span class="tag tag-pdf">{val}</span></td></tr>\n'
        html += _render_pdf_rows(truly_missing, _fmt_entry_parts, _missing_excel_cols, 4)
        html += """      </table>
    </div>
  </div>
"""

    # --- Field exists in Excel but value is None ---
    if has_field_none:
        html += f"""
  <div class="section">
    <div class="section-header collapsible" style="background: #9b59b6;" onclick="toggleSection(this)">
      <span><span class="arrow">&#9660;</span>In PDF, Excel field exists but value is None</span>
      <span class="badge">{len(has_field_none)}</span>
    </div>
    <div class="section-body">
      <table>
        <tr><th>#</th><th>Subsection</th><th>Permission</th><th>PDF Value</th><th>Excel Row</th></tr>
"""
        def _none_cols(i, entry, sub, fld):
            val = _html_escape(entry["permissions_str"])
            row = entry.get("_excel_row", "?")
            return f'        <tr><td>{i}</td><td>{_html_escape(sub)}</td><td>{_html_escape(fld)}</td><td><span class="tag tag-pdf">{val}</span></td><td>{row}</td></tr>\n'
        html += _render_pdf_rows(has_field_none, _fmt_entry_parts, _none_cols, 5)
        html += """      </table>
    </div>
  </div>
"""

    # --- In Excel but not in PDF ---
    if missing_in_pdf:
        html += f"""
  <div class="section">
    <div class="section-header collapsible collapsed" style="background: #3498db;" onclick="toggleSection(this)">
      <span><span class="arrow">&#9660;</span>In Excel but NOT FOUND in PDF</span>
      <span class="badge">{len(missing_in_pdf)}</span>
    </div>
    <div class="section-body hidden-body">
      <table>
        <tr><th>#</th><th>Grouping</th><th>Permission</th><th>Excel Value</th><th>Excel Row</th></tr>
"""
        last_elem = None
        for i, entry in enumerate(missing_in_pdf, 1):
            elem = entry["element"]
            if elem != last_elem:
                html += f'        <tr class="group-row"><td colspan="5">{_html_escape(elem)}</td></tr>\n'
                last_elem = elem
            html += f'        <tr><td>{i}</td><td>{_html_escape(entry["grouping"])}</td><td>{_html_escape(entry["permission"])}</td><td><span class="tag tag-excel">{_html_escape(entry["value"])}</span></td><td>{entry["row"]}</td></tr>\n'
        html += """      </table>
    </div>
  </div>
"""

    html += """</div>
<script>
const filterInput = document.getElementById('filterInput');
const filterInfo = document.getElementById('filterInfo');

function applyFilter() {
  const term = filterInput.value.trim().toLowerCase();
  const rows = document.querySelectorAll('.section-body table tr:not(:first-child)');
  let hidden = 0, total = 0;
  rows.forEach(row => {
    total++;
    const text = row.textContent.toLowerCase();
    if (term && !text.includes(term)) {
      row.classList.add('hidden-row');
      hidden++;
    } else {
      row.classList.remove('hidden-row');
    }
  });
  document.querySelectorAll('.section').forEach(sec => {
    const tbl = sec.querySelector('table');
    if (!tbl) return;
    const visible = tbl.querySelectorAll('tr:not(:first-child):not(.hidden-row)').length;
    const badge = sec.querySelector('.badge');
    if (badge) badge.textContent = visible;
  });
  filterInfo.textContent = term ? `Showing ${total - hidden} of ${total} rows` : '';
}

filterInput.addEventListener('input', applyFilter);

function clearFilter() {
  filterInput.value = '';
  applyFilter();
}

function toggleSection(header) {
  header.classList.toggle('collapsed');
  const body = header.nextElementSibling;
  if (body) body.classList.toggle('hidden-body');
}
</script>
</body>
</html>"""

    return html


# --------------------------------------------------------------------------
# PDF TO EXCEL EXPORT
# --------------------------------------------------------------------------

def generate_pdf_excel(pdf_entries, output_path):
    """
    Export the extracted PDF permissions to an Excel file.
    Columns: Subsection, Permission, Value
    Section changes are shown as merged separator rows with a background colour.
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PDF Permissions"

    headers = ["Subsection", "Permission", "Value"]
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    thin_border = Border(bottom=Side(style="thin", color="DDDDDD"))
    section_font = Font(bold=True, color="FFFFFF", size=11)
    section_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    row_idx = 2
    last_section = None

    for entry in pdf_entries:
        section = entry.get("element") or entry["section"]
        if section != last_section:
            # Insert a section separator row
            cell = ws.cell(row=row_idx, column=1, value=section)
            cell.font = section_font
            cell.fill = section_fill
            for col in range(2, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = section_fill
            ws.merge_cells(start_row=row_idx, start_column=1,
                           end_row=row_idx, end_column=len(headers))
            last_section = section
            row_idx += 1

        subsection = entry.get("subsection", "")
        if subsection == section:
            subsection = ""
        ws.cell(row=row_idx, column=1, value=subsection)
        ws.cell(row=row_idx, column=2, value=entry["field"])
        ws.cell(row=row_idx, column=3, value=entry["permissions_str"])
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).border = thin_border
        row_idx += 1

    # Auto-fit column widths
    for col in range(1, len(headers) + 1):
        max_len = len(headers[col - 1])
        for r in range(2, min(row_idx, 200)):
            val = ws.cell(row=r, column=col).value
            if val and len(str(val)) > max_len:
                max_len = len(str(val))
        ws.column_dimensions[chr(64 + col)].width = min(max_len + 4, 60)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:C{row_idx - 1}"

    wb.save(output_path)
    wb.close()
    return output_path


# --------------------------------------------------------------------------
# MAIN
# --------------------------------------------------------------------------

def main():
    print("=" * 70)
    print("  PDF vs Excel Permission Comparator")
    print("  Role: ALL_MGR_GL_Manager")
    print("=" * 70)

    # Step 1: Extract PDF permissions
    print("\n[1/4] Extracting permissions from PDF...")
    pdf_entries = extract_pdf_permissions(PDF_PATH)
    print(f"      Found {len(pdf_entries)} permission entries in PDF.")

    # Step 2: Extract Excel permissions
    print("\n[2/4] Reading Excel workbook...")
    excel_entries = extract_excel_permissions(EXCEL_PATH, EXCEL_SHEET, ROLE_COLUMN)
    print(f"      Found {len(excel_entries)} rows in Excel for ALL_MGR_GL_Manager.")
    active_excel = [e for e in excel_entries if e["value"] not in ("None", "None (not in system)")]
    print(f"      Of which {len(active_excel)} have actual permission values.")

    # Step 3: Compare
    print("\n[3/4] Comparing PDF vs Excel...")
    missing_in_excel, missing_in_pdf, mismatches, matched = compare_pdf_vs_excel(
        pdf_entries, excel_entries
    )

    # Step 4: Reports
    print("\n[4/5] Generating reports...")
    report = generate_report(missing_in_excel, missing_in_pdf, mismatches, matched,
                             total_pdf=len(pdf_entries),
                             total_excel=len(excel_entries),
                             total_excel_active=len(active_excel))

    output_txt = os.path.join(SCRIPT_DIR, "pdf_vs_excel_comparison_report.txt")
    with open(output_txt, "w", encoding="utf-8") as f:
        f.write(report)

    output_html = os.path.join(SCRIPT_DIR, "pdf_vs_excel_comparison_report.html")
    html = generate_html_report(missing_in_excel, missing_in_pdf, mismatches, matched,
                                total_pdf=len(pdf_entries),
                                total_excel=len(excel_entries),
                                total_excel_active=len(active_excel))
    with open(output_html, "w", encoding="utf-8") as f:
        f.write(html)

    # Step 5: PDF to Excel export
    print("\n[5/5] Exporting PDF permissions to Excel...")
    output_xlsx = os.path.join(SCRIPT_DIR, "pdf_permissions_export.xlsx")
    generate_pdf_excel(pdf_entries, output_xlsx)

    print(report)
    print(f"\n[DONE] Reports saved:")
    print(f"  TXT:   {os.path.basename(output_txt)}")
    print(f"  HTML:  {os.path.basename(output_html)}")
    print(f"  Excel: {os.path.basename(output_xlsx)}")


if __name__ == "__main__":
    main()
